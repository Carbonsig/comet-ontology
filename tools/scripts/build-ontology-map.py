#!/usr/bin/env python3
"""Build the consolidated COMET ontology map artifacts.

Single source-of-truth extractor. Reads every place a COMET term is defined —
the JSON-LD context, the JSON Schemas, the OWL/TTL extension modules and the
alignment crosswalks — normalises them into one term table plus a node/edge
graph, then emits:

  docs/ontology-data.json        Consolidated graph + term table (feeds the
                                 interactive schema-map.html page).
  docs/comet-ontology-values.xlsx  Multi-sheet workbook: Terms, Namespaces,
                                 Alignments, Schema Fields, Layers, Summary.

Run:
    tools/.venv/bin/python tools/scripts/build-ontology-map.py

Dependencies: rdflib, openpyxl  (see tools/requirements.txt).
"""

from __future__ import annotations

import html as _html
import json
import re
import sys
from collections import Counter, defaultdict
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from rdflib import Graph, RDF, RDFS, OWL, Namespace, URIRef, Literal
from rdflib.namespace import SKOS, DCTERMS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DOCS = ROOT / "docs"
CONTEXT_FILE = ROOT / "comet-context.jsonld"
SCHEMA_DIR = ROOT / "tools" / "schemas"
EXT_DIR = ROOT / "ext"
GLOSSARY_FILE = DOCS / "glossary.html"
CAD_TRUST_XLSX = DOCS / "CAD-Trust-Data-Dictionary-v2.0.2.xlsx"

PROV = Namespace("http://www.w3.org/ns/prov#")

# Namespace prefix → IRI base. Mirrors comet-context.jsonld.
NS_BASES: dict[str, str] = {
    "comet": "https://comet.carbon/v1/core#",
    "comet-rs": "https://comet.carbon/ext/responsiblesteel#",
    "comet-cn": "https://comet.carbon/ext/iso14068#",
    "comet-pcf": "https://comet.carbon/v1/pcf#",
    "comet-eac": "https://comet.carbon/v1/eac#",
    "comet-ver": "https://comet.carbon/v1/ver#",
    "comet-mkt": "https://comet.carbon/v1/market#",
    "comet-sc": "https://comet.carbon/v1/supplychain#",
    "comet-ef": "https://comet.carbon/v1/emfactor#",
    # The RS extension TTLs also use this alternate base in alignments.
    "comet-rs-alt": "https://ontology.materialintelligence.ai/comet/responsiblesteel/",
}

# Human-readable namespace metadata (the COMET seven-layer stack + extensions).
NS_META: dict[str, dict[str, str]] = {
    "comet": {"name": "Core", "layer": "L1 · Core", "color": "#1a3a6b"},
    "comet-pcf": {"name": "Product Carbon Footprint", "layer": "L4 · PCF", "color": "#1a6b3c"},
    "comet-eac": {"name": "Environmental Attribute Certificate", "layer": "L5 · EAC", "color": "#9a7c2f"},
    "comet-ver": {"name": "Verification", "layer": "L6 · Verification", "color": "#1a5f6b"},
    "comet-mkt": {"name": "Market", "layer": "L7 · Market", "color": "#7a2f6b"},
    "comet-sc": {"name": "Supply Chain", "layer": "L3 · Supply Chain", "color": "#6b4a1a"},
    "comet-ef": {"name": "Emission Factor", "layer": "L2 · Emission Factor", "color": "#3c1a6b"},
    "comet-rs": {"name": "Responsible Steel (ext)", "layer": "Extension", "color": "#c8360a"},
    "comet-cn": {"name": "ISO 14068 Carbon Neutrality (ext)", "layer": "Extension", "color": "#0a7c8c"},
    "cadtrust": {"name": "CAD Trust Data Dictionary", "layer": "CAD Trust (incorporated)", "color": "#5a5a5a", "graph_default": False},
    "cadpick": {"name": "CAD Trust Picklist Values", "layer": "CAD Trust (incorporated)", "color": "#8a8a8a", "graph_default": False},
}

# glossary.html section id → COMET namespace prefix (the authoritative
# seven-layer vocabulary lives in these tables, not in TTL here).
GLOSSARY_SECTIONS: dict[str, str] = {
    "l1-core": "comet",
    "l2-emfactor": "comet-ef",
    "l3-supplychain": "comet-sc",
    "l4-pcf": "comet-pcf",
    "l5-eac": "comet-eac",
    "l6-verification": "comet-ver",
    "l7-market": "comet-mkt",
}

# Standards-Alignment cell prefixes → friendly standard name (glossary column).
ALIGN_NAME_MAP: dict[str, str] = {
    "cad trust": "CAD Trust", "pact": "WBCSD PACT", "schema": "schema.org",
    "ghg protocol": "GHG Protocol", "iso 14064": "ISO 14064", "iso 14067": "ISO 14067",
    "iso 14068": "ISO 14068-1", "cbam": "EU CBAM", "esrs": "EU ESRS", "esg": "ESG",
    "iso 14040": "ISO 14040", "iso 14044": "ISO 14044", "prov": "W3C PROV",
    "qudt": "QUDT", "geosparql": "GeoSPARQL", "iec": "IEC", "irec": "I-REC",
    "gs": "Gold Standard", "verra": "Verra", "vcs": "Verra VCS",
    # v0.3.0 carbon-verification-market standards (friendly-name crosswalks; no
    # invented external IRIs — target_iri stays empty until a citable scheme exists).
    "ghg protocol scope 3": "GHG Protocol Scope 3", "ghg protocol: scope 1": "GHG Protocol",
    "ghg protocol: scope 2": "GHG Protocol",
    "iso 14064-3": "ISO 14064-3", "iso 14065": "ISO 14065",
    "isae 3410": "ISAE 3410", "isae 3000": "ISAE 3000",
    "issb s2": "ISSB S2", "ifrs s2": "ISSB S2", "esrs e1-6": "EU ESRS",
    "eu ets mrv": "EU ETS MRV", "ira 45v": "IRA 45V", "45vh2-greet": "IRA 45V",
    "icao corsia": "ICAO CORSIA", "icao corsia (euc criteria)": "ICAO CORSIA",
    "sbti": "SBTi", "en 15804+a2": "EN 15804+A2", "pas 2050": "PAS 2050",
    "ca sb 253": "California SB 253", "ca sb 261": "California SB 261",
}

# IRI bases used by *external* vocabularies that COMET aligns to.
EXTERNAL_BASES: dict[str, str] = {
    "iso14068": "https://www.iso.org/standard/14068-1#",
    "iso14064": "https://www.iso.org/standard/14064-1#",
    "iso14067": "https://www.iso.org/standard/14067#",
    "cbam": "https://ontology.materialintelligence.ai/cbam/",
    "esrs": "https://data.europa.eu/api/hub/store/schema/eu.europa.ec.eupl.schema.EBA.esrs.1.0/",
    "ghgp": "https://ghgprotocol.org/scope/",
    "pact": "https://wbcsd.org/spec/pact/pathfinder/v3/",
    "sdg": "https://metadata.un.org/sdg/",
    "qudt-unit": "http://qudt.org/vocab/UNIT/",
    # CarbonSig Verifier Hub export contract (v4) — crosswalk target.
    "csig": "https://api.carbonsig.com/schemas/verifier-export/v4#",
}

# Friendly display names + colours for external standards (alignment targets).
EXTERNAL_META: dict[str, dict[str, str]] = {
    "iso14068": {"name": "ISO 14068-1", "color": "#555"},
    "iso14064": {"name": "ISO 14064", "color": "#666"},
    "iso14067": {"name": "ISO 14067", "color": "#777"},
    "cbam": {"name": "EU CBAM", "color": "#0a4ea0"},
    "esrs": {"name": "EU ESRS", "color": "#3a6ea0"},
    "ghgp": {"name": "GHG Protocol", "color": "#1a6b3c"},
    "pact": {"name": "WBCSD PACT", "color": "#9a7c2f"},
    "sdg": {"name": "UN SDG", "color": "#2a8ab0"},
    "qudt-unit": {"name": "QUDT Unit", "color": "#888"},
    "csig": {"name": "CarbonSig Verifier Export v4", "color": "#0a7c5a"},
}


def canonical_iri(iri: str) -> str:
    """Collapse the alternate comet-rs base onto the canonical one.

    The alignment TTLs declare comet-rs under
    ``…materialintelligence.ai/comet/responsiblesteel/`` while the module TTL
    uses ``comet.carbon/ext/responsiblesteel#``. Same logical terms — normalise
    so alignment edges connect to the defined term nodes.
    """
    alt = NS_BASES["comet-rs-alt"]
    if iri.startswith(alt):
        return NS_BASES["comet-rs"] + iri[len(alt):]
    return iri

# Crosswalk predicate → relationship label.
ALIGN_PREDICATES = {
    str(OWL.equivalentClass): "equivalentClass",
    str(OWL.equivalentProperty): "equivalentProperty",
    str(SKOS.exactMatch): "exactMatch",
    str(SKOS.closeMatch): "closeMatch",
    str(SKOS.relatedMatch): "relatedMatch",
    str(SKOS.broadMatch): "broadMatch",
    str(SKOS.narrowMatch): "narrowMatch",
    str(SKOS.related): "related",
}

# Type IRI → friendly kind label.
KIND_LABELS = {
    str(OWL.Class): "Class",
    str(RDFS.Class): "Class",
    str(OWL.ObjectProperty): "ObjectProperty",
    str(OWL.DatatypeProperty): "DatatypeProperty",
    str(OWL.AnnotationProperty): "AnnotationProperty",
    str(OWL.NamedIndividual): "Individual",
    str(RDF.Property): "Property",
}


def prefix_for(iri: str) -> tuple[str | None, str]:
    """Return (prefix, local_name) for an IRI, or (None, iri) if unknown."""
    for prefix, base in {**NS_BASES, **EXTERNAL_BASES}.items():
        if iri.startswith(base):
            if prefix == "comet-rs-alt":
                prefix = "comet-rs"
            return prefix, iri[len(base):]
    # Fall back: split on # or last /
    for sep in ("#", "/"):
        if sep in iri:
            return None, iri.rsplit(sep, 1)[-1]
    return None, iri


def en_value(graph: Graph, subj: URIRef, *preds) -> str | None:
    """First English (or untagged) literal across the given predicates."""
    for pred in preds:
        # Prefer an explicit @en literal.
        for obj in graph.objects(subj, pred):
            if isinstance(obj, Literal) and obj.language == "en":
                return str(obj)
        for obj in graph.objects(subj, pred):
            if isinstance(obj, Literal) and obj.language in (None, ""):
                return str(obj)
        for obj in graph.objects(subj, pred):
            if isinstance(obj, Literal):
                return str(obj)
    return None


def all_langs(graph: Graph, subj: URIRef, *preds) -> dict[str, str]:
    """Map language-code → label across the given predicates."""
    out: dict[str, str] = {}
    for pred in preds:
        for obj in graph.objects(subj, pred):
            if isinstance(obj, Literal) and obj.language:
                out.setdefault(obj.language, str(obj))
    return out


def load_ttl_graph() -> tuple[Graph, list[Path], list[Path]]:
    """Load all extension + alignment TTLs into one rdflib graph."""
    graph = Graph()
    module_files = sorted(EXT_DIR.glob("*/comet-ext-*.ttl"))
    align_files = sorted(EXT_DIR.glob("*/alignments/*.ttl"))
    label_files = sorted(EXT_DIR.glob("*/labels/*.ttl"))
    for path in module_files + align_files + label_files:
        try:
            graph.parse(str(path), format="turtle")
        except Exception as exc:  # noqa: BLE001 - report and continue
            print(f"  ! failed to parse {path.relative_to(ROOT)}: {exc}", file=sys.stderr)
    return graph, module_files, align_files


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------


def extract_from_ttl(graph: Graph) -> tuple[dict[str, dict], list[dict]]:
    """Pull terms + structural/alignment edges from the merged TTL graph."""
    terms: dict[str, dict] = {}
    edges: list[dict] = []

    # Subjects that carry an rdf:type we care about.
    for subj in set(graph.subjects(RDF.type, None)):
        if not isinstance(subj, URIRef):
            continue
        types = {str(t) for t in graph.objects(subj, RDF.type)}
        kinds = [KIND_LABELS[t] for t in types if t in KIND_LABELS]
        if not kinds:
            continue
        iri = canonical_iri(str(subj))
        prefix, local = prefix_for(iri)
        if prefix not in NS_META:  # only COMET-owned terms become first-class rows
            continue
        kind = kinds[0]
        label = en_value(graph, subj, RDFS.label, SKOS.prefLabel) or local
        definition = en_value(graph, subj, SKOS.definition, RDFS.comment)
        parents = [prefix_for(str(o))[1] for o in graph.objects(subj, RDFS.subClassOf)
                   if isinstance(o, URIRef)]
        domains = [prefix_for(str(o))[1] for o in graph.objects(subj, RDFS.domain)
                   if isinstance(o, URIRef)]
        ranges = [prefix_for(str(o))[1] for o in graph.objects(subj, RDFS.range)
                  if isinstance(o, URIRef)]
        langs = all_langs(graph, subj, RDFS.label, SKOS.prefLabel)

        rec = terms.setdefault(iri, {
            "iri": iri, "curie": f"{prefix}:{local}", "prefix": prefix,
            "local": local, "kind": kind,
            "namespace": NS_META[prefix]["name"], "layer": NS_META[prefix]["layer"],
            "source": "ttl",
        })
        rec.update({
            "label": label, "definition": definition,
            "subClassOf": sorted(set(parents)), "domain": sorted(set(domains)),
            "range": sorted(set(ranges)), "languages": sorted(langs.keys()),
        })

        # Structural edges (subClassOf / domain / range).
        for parent in graph.objects(subj, RDFS.subClassOf):
            if isinstance(parent, URIRef):
                edges.append({"from": iri, "to": canonical_iri(str(parent)),
                              "rel": "subClassOf"})
        for rng in graph.objects(subj, RDFS.range):
            tgt = canonical_iri(str(rng))
            if isinstance(rng, URIRef) and prefix_for(tgt)[0] in NS_META:
                edges.append({"from": iri, "to": tgt, "rel": "range"})

    # Alignment edges (any subject, including ones without rdf:type rows).
    for pred_iri, rel in ALIGN_PREDICATES.items():
        pred = URIRef(pred_iri)
        for s, o in graph.subject_objects(pred):
            if not (isinstance(s, URIRef) and isinstance(o, URIRef)):
                continue
            s_iri, o_iri = canonical_iri(str(s)), canonical_iri(str(o))
            if prefix_for(s_iri)[0] not in NS_META and prefix_for(o_iri)[0] not in NS_META:
                continue
            edges.append({"from": s_iri, "to": o_iri, "rel": rel})

    return terms, edges


def extract_alignments(graph: Graph) -> list[dict]:
    """Flat crosswalk rows: COMET term → external standard term."""
    rows: list[dict] = []
    for pred_iri, rel in ALIGN_PREDICATES.items():
        pred = URIRef(pred_iri)
        for s, o in graph.subject_objects(pred):
            if not (isinstance(s, URIRef) and isinstance(o, URIRef)):
                continue
            s_prefix, s_local = prefix_for(canonical_iri(str(s)))
            o_prefix, o_local = prefix_for(canonical_iri(str(o)))
            comet_side, ext_side = None, None
            if s_prefix in NS_META:
                comet_side, ext_side = (s_prefix, s_local), (o_prefix, o_local, str(o))
            elif o_prefix in NS_META:
                comet_side, ext_side = (o_prefix, o_local), (s_prefix, s_local, str(s))
            else:
                continue
            note = (en_value(graph, s, SKOS.note, RDFS.comment)
                    or en_value(graph, o, SKOS.note, RDFS.comment))
            std = ext_side[0]
            std_name = EXTERNAL_META.get(std, {}).get("name", std or "external")
            rows.append({
                "comet_curie": f"{comet_side[0]}:{comet_side[1]}",
                "comet_local": comet_side[1],
                "relation": rel,
                "target_standard": std_name,
                "target_prefix": std or "external",
                "target_term": ext_side[1],
                "target_iri": ext_side[2],
                "note": note or "",
            })
    # De-dup identical rows.
    seen, out = set(), []
    for r in rows:
        key = (r["comet_curie"], r["relation"], r["target_iri"])
        if key not in seen:
            seen.add(key)
            out.append(r)
    return sorted(out, key=lambda r: (r["comet_curie"], r["relation"]))


class _GlossaryParser(HTMLParser):
    """Collect (section_id, [(td_class, text), …]) rows from glossary.html."""

    def __init__(self) -> None:
        super().__init__()
        self.section: str | None = None
        self.in_tbody = self.in_tr = self.in_td = False
        self.cur_class = ""
        self.cur_text: list[str] = []
        self.cells: list[tuple[str, str]] = []
        self.rows: list[tuple[str, list[tuple[str, str]]]] = []

    def handle_starttag(self, tag: str, attrs):
        a = dict(attrs)
        if tag == "section" and a.get("id") in GLOSSARY_SECTIONS:
            self.section = a["id"]
        elif tag == "tbody":
            self.in_tbody = True
        elif tag == "tr" and self.in_tbody:
            self.in_tr, self.cells = True, []
        elif tag == "td" and self.in_tr:
            self.in_td, self.cur_class, self.cur_text = True, a.get("class", ""), []

    def handle_endtag(self, tag: str):
        if tag == "td" and self.in_td:
            self.in_td = False
            self.cells.append((self.cur_class, "".join(self.cur_text).strip()))
        elif tag == "tr" and self.in_tr:
            self.in_tr = False
            if self.cells and self.section in GLOSSARY_SECTIONS:
                self.rows.append((self.section, self.cells))
        elif tag == "tbody":
            self.in_tbody = False
        elif tag == "section":
            self.section = None

    def handle_data(self, data: str):
        if self.in_td:
            self.cur_text.append(data)

    def handle_entityref(self, name: str):
        if self.in_td:
            self.cur_text.append(_html.unescape(f"&{name};"))


def extract_from_glossary() -> tuple[dict[str, dict], list[dict]]:
    """Parse the COMET seven-layer vocabulary tables in glossary.html.

    Returns (terms_by_curie, alignment_rows). This is the bulk of the core
    (non-extension) ontology — it is not shipped as TTL in this repo.
    """
    if not GLOSSARY_FILE.exists():
        return {}, []
    p = _GlossaryParser()
    p.feed(GLOSSARY_FILE.read_text(encoding="utf-8"))
    terms: dict[str, dict] = {}
    aligns: list[dict] = []

    for section, cells in p.rows:
        prefix = GLOSSARY_SECTIONS[section]
        name = cells[0][1].strip()
        if not name or name in ("—", "—"):
            continue
        kind, datatype, align_txt, desc = "Property", "", "", ""
        for i, (cls, txt) in enumerate(cells):
            if "td-type" in cls:
                kind = "Class" if txt.strip().lower().startswith("class") else "Property"
            elif "td-datatype" in cls:
                datatype = "" if txt in ("—", "—") else txt
            elif "td-align" in cls:
                align_txt = "" if txt in ("—", "—") else txt
            elif i != 0 and not cls and txt and txt not in ("—", "—"):
                desc = desc or txt
        curie = name if ":" in name else f"{prefix}:{name}"
        pfx = curie.split(":", 1)[0]
        local = curie.split(":", 1)[1]
        meta = NS_META.get(pfx, NS_META[prefix])
        terms[curie] = {
            "iri": curie, "curie": curie, "prefix": pfx if pfx in NS_META else prefix,
            "local": local, "label": local.split(".")[-1], "kind": kind,
            "namespace": meta["name"], "layer": meta["layer"],
            "datatype": datatype, "definition": desc, "source": "glossary",
        }
        # Standards-alignment column → crosswalk rows (one per target).
        # Two accepted chunk forms:
        #   "<standard>: <target term>"  → head matched against ALIGN_NAME_MAP
        #   "<standard>"  (bare name)    → whole chunk matched; target term blank
        # Bare names let coverage register for standards without a per-term
        # external identifier (ISSB S2, SBTi, …) without inventing IRIs.
        for chunk in re.split(r"\s*;\s*", align_txt):
            chunk = chunk.strip()
            if not chunk:
                continue
            if ":" in chunk:
                head, tail = chunk.split(":", 1)
                head, tail = head.strip(), tail.strip()
            else:
                head, tail = chunk, ""
            std = ALIGN_NAME_MAP.get(head.lower())
            if not std:
                continue
            aligns.append({
                "comet_curie": curie, "comet_local": local, "relation": "mapsTo",
                "target_standard": std, "target_prefix": head.lower(),
                "target_term": tail, "target_iri": "", "note": "",
            })
    return terms, aligns


def extract_from_cad_trust() -> dict[str, dict]:
    """Parse the CAD Trust Data Dictionary xlsx (incorporated standard).

    Pulls real data-field rows from the 'Data Fields' sheet and picklist
    option values from 'Picklist Values'. Group-header / blank rows skipped.
    """
    if not CAD_TRUST_XLSX.exists():
        return {}
    terms: dict[str, dict] = {}
    wb = load_workbook(CAD_TRUST_XLSX, read_only=True, data_only=True)

    # --- Data Fields: r1.. header at row 6, data from row 7 ---
    if "Data Fields" in wb.sheetnames:
        ws = wb["Data Fields"]
        for r in ws.iter_rows(min_row=7, values_only=True):
            table = (r[1] or "").strip() if len(r) > 1 and r[1] else ""
            api = (r[3] or "").strip() if len(r) > 3 and r[3] else ""
            if not api:  # group-header / blank row
                continue
            dtype = (r[4] or "").strip() if len(r) > 4 and r[4] else ""
            desc = (r[7] or "").strip() if len(r) > 7 and r[7] else ""
            curie = f"cadtrust:{table}.{api}" if table else f"cadtrust:{api}"
            terms[curie] = {
                "iri": curie, "curie": curie, "prefix": "cadtrust",
                "local": f"{table}.{api}" if table else api, "label": api,
                "kind": "Field", "namespace": NS_META["cadtrust"]["name"],
                "layer": NS_META["cadtrust"]["layer"], "datatype": dtype,
                "definition": desc, "source": "cad-trust",
            }

    # --- Picklist Values: column-per-picklist, option values down rows ---
    if "Picklist Values" in wb.sheetnames:
        ws = wb["Picklist Values"]
        grid = list(ws.iter_rows(min_row=7, values_only=True))
        # Row 8 (index 1 here) holds picklist names; row 9 descriptions; 10+ values.
        if len(grid) > 3:
            names = grid[1]
            for col in range(1, len(names)):
                pname = (names[col] or "").strip() if names[col] else ""
                if not pname:
                    continue
                pname = pname.split("\n")[0].strip()
                for row in grid[3:]:
                    val = (row[col] or "").strip() if col < len(row) and row[col] else ""
                    if not val:
                        continue
                    curie = f"cadpick:{pname}.{val}"[:160]
                    terms[curie] = {
                        "iri": curie, "curie": curie, "prefix": "cadpick",
                        "local": f"{pname} = {val}", "label": val,
                        "kind": "PicklistValue", "namespace": NS_META["cadpick"]["name"],
                        "layer": NS_META["cadpick"]["layer"], "datatype": "enum",
                        "definition": f"Picklist option for {pname}.", "source": "cad-trust",
                    }
    wb.close()
    return terms


def extract_from_context() -> dict[str, dict]:
    """Terms declared in comet-context.jsonld (adds datatype hints)."""
    ctx = json.loads(CONTEXT_FILE.read_text())["@context"]
    terms: dict[str, dict] = {}
    for key, val in ctx.items():
        if key.startswith("@") or ":" not in str(val) and not isinstance(val, dict):
            continue
        if isinstance(val, dict):
            iri = val.get("@id", "")
            datatype = val.get("@type", "")
        else:
            iri = val
            datatype = ""
        if not isinstance(iri, str) or ":" not in iri:
            continue
        prefix = iri.split(":", 1)[0]
        if prefix not in NS_META:
            continue
        local = iri.split(":", 1)[1]
        base = NS_BASES.get(prefix, "")
        full = base + local
        kind = "Property" if (datatype or key[:1].islower()) else "Class"
        terms[full] = {
            "iri": full, "curie": iri, "prefix": prefix, "local": local,
            "term": key, "kind": kind,
            "namespace": NS_META[prefix]["name"], "layer": NS_META[prefix]["layer"],
            "datatype": datatype.replace("xsd:", "") if datatype else "",
            "source": "context",
        }
    return terms


def extract_from_schemas() -> list[dict]:
    """Flatten the JSON Schemas into a field table."""
    rows: list[dict] = []
    for schema_path in sorted(SCHEMA_DIR.glob("*.schema.json")):
        data = json.loads(schema_path.read_text())
        schema_name = data.get("title", schema_path.stem)
        module = schema_path.stem.replace("comet-", "").replace(".schema", "")

        def walk(defn: dict, entity: str) -> None:
            props = defn.get("properties", {})
            required = set(defn.get("required", []))
            for field, spec in props.items():
                if field.startswith("@"):
                    continue
                ftype = spec.get("type", "")
                if isinstance(ftype, list):
                    ftype = "/".join(ftype)
                enum = spec.get("enum")
                rows.append({
                    "schema": schema_name,
                    "module": module,
                    "entity": entity,
                    "field": field,
                    "type": ftype or ("enum" if enum else ""),
                    "required": field in required,
                    "enum": ", ".join(map(str, enum)) if enum else "",
                    "description": spec.get("description", ""),
                })

        for entity, defn in data.get("$defs", {}).items():
            if isinstance(defn, dict):
                walk(defn, entity)
        if "properties" in data:
            walk(data, data.get("title", "root"))
    return rows


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------


def merge_terms(*sources: dict) -> list[dict]:
    """Merge term records from every source, keyed by CURIE (fallback IRI).

    Sources are applied in order; later sources fill blanks left by earlier
    ones but do not clobber non-empty values. Provenance is accumulated.
    Earlier sources should be the more descriptive ones (glossary, TTL).
    """
    merged: dict[str, dict] = {}
    for src in sources:
        for rec in src.values():
            key = rec.get("curie") or rec.get("iri")
            if key not in merged:
                merged[key] = dict(rec)
                continue
            base = merged[key]
            for k, v in rec.items():
                if k == "source":
                    continue
                if base.get(k) in (None, "", [], {}) and v not in (None, "", [], {}):
                    base[k] = v
            srcs = set()
            for s in (base.get("source"), rec.get("source")):
                if s:
                    srcs.update(s.split("+"))
            base["source"] = "+".join(sorted(srcs))
    out = []
    for rec in merged.values():
        rec.setdefault("label", rec.get("local", ""))
        rec.setdefault("definition", "")
        rec.setdefault("datatype", "")
        for f in ("subClassOf", "domain", "range", "languages"):
            rec.setdefault(f, [])
        out.append(rec)
    return sorted(out, key=lambda r: (r["prefix"], r["kind"], r["local"].lower()))


def build_graph_payload(terms: list[dict], edges: list[dict]) -> dict:
    """Build cytoscape-friendly nodes/edges.

    Defined terms become full nodes. Endpoints referenced by an edge but not
    defined in-repo are synthesised as lightweight stub nodes: COMET-namespace
    references (core/pcf/eac/sc … layers not shipped as TTL here) and external
    standard targets (ISO, CBAM, ESRS, GHG Protocol, PACT). This keeps the
    graph connected across the full seven-layer stack and the crosswalks.
    """
    # The flat CAD Trust dictionary (≈480 fields/picklists) has no relational
    # structure — it belongs in the table + XLSX, not the relationship graph.
    GRAPH_EXCLUDE = {"cadtrust", "cadpick"}
    nodes_by_id: dict[str, dict] = {}
    by_curie: dict[str, str] = {}
    for t in terms:
        if t["prefix"] in GRAPH_EXCLUDE:
            continue
        nodes_by_id[t["iri"]] = {
            "id": t["iri"],
            "label": t["label"] or t["local"],
            "curie": t["curie"],
            "kind": t["kind"],
            "prefix": t["prefix"],
            "namespace": t["namespace"],
            "layer": t["layer"],
            "color": NS_META.get(t["prefix"], {}).get("color", "#888"),
            "definition": (t.get("definition") or "")[:400],
            "defined": True,
        }
        by_curie[t["curie"]] = t["iri"]

    # Synthesise partOf edges for dotted core properties (X.y → X) so the
    # glossary vocabulary connects its properties to their owning class.
    part_edges: list[dict] = []
    for t in terms:
        if t["prefix"] in GRAPH_EXCLUDE or "." not in t.get("local", ""):
            continue
        parent_curie = f"{t['prefix']}:{t['local'].split('.')[0]}"
        parent_iri = by_curie.get(parent_curie)
        if parent_iri and parent_iri != t["iri"]:
            part_edges.append({"from": t["iri"], "to": parent_iri, "rel": "partOf"})
    edges = edges + part_edges

    def ensure_stub(iri: str) -> bool:
        """Add a stub node for a referenced-but-undefined IRI. Returns kept?"""
        if iri in nodes_by_id:
            return True
        prefix, local = prefix_for(iri)
        if prefix in NS_META:  # COMET layer term defined elsewhere
            meta = NS_META[prefix]
            nodes_by_id[iri] = {
                "id": iri, "label": local, "curie": f"{prefix}:{local}",
                "kind": "Reference", "prefix": prefix, "namespace": meta["name"],
                "layer": meta["layer"], "color": meta["color"],
                "definition": "", "defined": False,
            }
            return True
        if prefix in EXTERNAL_META:  # external standard target
            meta = EXTERNAL_META[prefix]
            nodes_by_id[iri] = {
                "id": iri, "label": local, "curie": f"{prefix}:{local}",
                "kind": "External", "prefix": prefix, "namespace": meta["name"],
                "layer": "External Standard", "color": meta["color"],
                "definition": "", "defined": False,
            }
            return True
        return False

    clean_edges, seen = [], set()
    for e in edges:
        if not (ensure_stub(e["from"]) and ensure_stub(e["to"])):
            continue
        key = (e["from"], e["to"], e["rel"])
        if key not in seen:
            seen.add(key)
            clean_edges.append(e)
    return {"nodes": list(nodes_by_id.values()), "edges": clean_edges}


def write_json(terms, edges, aligns, schema_rows, ns_summary) -> Path:
    graph = build_graph_payload(terms, edges)
    payload = {
        "generated": "build-ontology-map.py",
        "namespaces": [
            {"prefix": p, **meta, "graph_default": meta.get("graph_default", True),
             "term_count": ns_summary.get(p, 0)}
            for p, meta in NS_META.items()
        ],
        "terms": terms,
        "graph": graph,
        "alignments": aligns,
        "schema_fields": schema_rows,
        "stats": {
            "terms": len(terms),
            "classes": sum(1 for t in terms if t["kind"] == "Class"),
            "properties": sum(1 for t in terms if "Property" in t["kind"]),
            "fields": sum(1 for t in terms if t["kind"] in ("Field", "PicklistValue")),
            "alignments": len(aligns),
            "schema_fields": len(schema_rows),
            "namespaces": sum(1 for n in NS_META.values() if True),
            "graph_nodes": len(graph["nodes"]),
            "graph_edges": len(graph["edges"]),
        },
    }
    out = DOCS / "ontology-data.json"
    body = json.dumps(payload, indent=2, ensure_ascii=False)
    out.write_text(body)
    # JS wrapper so schema-map.html works when opened directly via file://
    # (browsers block fetch() of local JSON; a <script src> still loads).
    (DOCS / "ontology-data.js").write_text(
        "// Auto-generated by build-ontology-map.py — do not edit.\n"
        "window.COMET_ONTOLOGY = " + body + ";\n"
    )
    return out


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

HEAD_FILL = PatternFill("solid", fgColor="0F0F0F")
HEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="D0CFC9")
BORDER = Border(bottom=THIN)


def _style_sheet(ws, widths: list[int], wrap_cols: set[int]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{ws.max_row}"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP if cell.column in wrap_cols else TOP
            cell.border = BORDER


def write_xlsx(terms, aligns, schema_rows, ns_summary) -> Path:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["COMET Ontology — Value Dump"])
    ws["A1"].font = Font(bold=True, size=16, name="Calibri")
    ws.append([])
    ws.append(["Metric", "Count"])
    counts = [
        ("Total terms", len(terms)),
        ("Classes", sum(1 for t in terms if t["kind"] == "Class")),
        ("Properties", sum(1 for t in terms if "Property" in t["kind"])),
        ("Individuals", sum(1 for t in terms if t["kind"] == "Individual")),
        ("CAD Trust data fields", sum(1 for t in terms if t["kind"] == "Field")),
        ("CAD Trust picklist values", sum(1 for t in terms if t["kind"] == "PicklistValue")),
        ("Namespaces / layers", len(NS_META)),
        ("Alignment crosswalks", len(aligns)),
        ("JSON Schema fields", len(schema_rows)),
        ("— terms from glossary (7 core layers)", sum(1 for t in terms if "glossary" in t.get("source", ""))),
        ("— terms from OWL/TTL extensions", sum(1 for t in terms if "ttl" in t.get("source", ""))),
        ("— terms from CAD Trust dictionary", sum(1 for t in terms if "cad-trust" in t.get("source", ""))),
    ]
    for label, val in counts:
        ws.append([label, val])
    for cell in ws[3]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12

    # --- Terms ---
    ws = wb.create_sheet("Terms")
    headers = ["CURIE", "Local Name", "Label", "Kind", "Namespace", "Layer",
               "Datatype", "subClassOf", "Domain", "Range", "Languages",
               "Source", "Definition", "IRI"]
    ws.append(headers)
    for t in terms:
        ws.append([
            t.get("curie", ""), t.get("local", ""), t.get("label", ""),
            t.get("kind", ""), t.get("namespace", ""), t.get("layer", ""),
            t.get("datatype", ""), ", ".join(t.get("subClassOf", [])),
            ", ".join(t.get("domain", [])), ", ".join(t.get("range", [])),
            ", ".join(t.get("languages", [])), t.get("source", ""),
            t.get("definition", ""), t.get("iri", ""),
        ])
    _style_sheet(ws, [24, 26, 30, 16, 22, 18, 12, 24, 20, 20, 14, 12, 60, 46],
                 wrap_cols={13})

    # --- Namespaces ---
    ws = wb.create_sheet("Namespaces")
    ws.append(["Prefix", "Name", "Layer", "IRI Base", "Term Count", "Color"])
    for p, meta in NS_META.items():
        ws.append([p, meta["name"], meta["layer"], NS_BASES.get(p, ""),
                   ns_summary.get(p, 0), meta["color"]])
    _style_sheet(ws, [14, 36, 22, 48, 12, 10], wrap_cols=set())

    # --- Alignments ---
    ws = wb.create_sheet("Alignments")
    ws.append(["COMET CURIE", "COMET Term", "Relation", "Target Standard",
               "Target Term", "Note", "Target IRI"])
    for a in aligns:
        ws.append([a["comet_curie"], a["comet_local"], a["relation"],
                   a["target_standard"], a["target_term"], a["note"],
                   a["target_iri"]])
    _style_sheet(ws, [26, 26, 18, 18, 28, 56, 46], wrap_cols={6})

    # --- Schema Fields ---
    ws = wb.create_sheet("Schema Fields")
    ws.append(["Schema", "Module", "Entity", "Field", "Type", "Required",
               "Enum Values", "Description"])
    for r in schema_rows:
        ws.append([r["schema"], r["module"], r["entity"], r["field"],
                   r["type"], "yes" if r["required"] else "", r["enum"],
                   r["description"]])
    _style_sheet(ws, [30, 12, 22, 24, 14, 10, 36, 60], wrap_cols={7, 8})

    out = DOCS / "comet-ontology-values.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    print("COMET ontology map builder")
    print("  loading TTL extensions + alignments …")
    graph, module_files, align_files = load_ttl_graph()
    print(f"    {len(graph)} triples from {len(module_files)} modules, "
          f"{len(align_files)} alignment files")

    ttl_terms, edges = extract_from_ttl(graph)
    ttl_aligns = extract_alignments(graph)

    print("  parsing glossary.html (7-layer core vocabulary) …")
    gloss_terms, gloss_aligns = extract_from_glossary()
    print(f"    {len(gloss_terms)} core terms, {len(gloss_aligns)} alignment rows")

    print("  parsing CAD Trust Data Dictionary xlsx …")
    cad_terms = extract_from_cad_trust()
    print(f"    {len(cad_terms)} CAD Trust fields + picklist values")

    ctx_terms = extract_from_context()
    schema_rows = extract_from_schemas()

    # Order = descriptiveness priority: glossary + TTL first, then context, CAD.
    terms = merge_terms(gloss_terms, ttl_terms, ctx_terms, cad_terms)

    # Combine + de-dup alignments from TTL crosswalks and the glossary column.
    seen_a, aligns = set(), []
    for a in ttl_aligns + gloss_aligns:
        key = (a["comet_curie"], a["relation"], a["target_standard"], a["target_term"])
        if key not in seen_a:
            seen_a.add(key)
            aligns.append(a)
    aligns.sort(key=lambda r: (r["comet_curie"], r["relation"]))

    ns_summary = Counter(t["prefix"] for t in terms)

    print(f"  TOTAL terms: {len(terms)}  | structural edges: {len(edges)}  | "
          f"alignments: {len(aligns)}  | schema fields: {len(schema_rows)}")
    print(f"  by namespace: {dict(ns_summary)}")

    json_path = write_json(terms, edges, aligns, schema_rows, ns_summary)
    xlsx_path = write_xlsx(terms, aligns, schema_rows, ns_summary)
    print(f"  wrote {json_path.relative_to(ROOT)}")
    print(f"  wrote {xlsx_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
